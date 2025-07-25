#!/usr/bin/env python3
"""
Create sample test files for manual vision testing.

Run this script to generate sample Excel and CSV files for testing the vision components.
"""

import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel files will not be created.")


def create_simple_table_xlsx(filepath: Path):
    """Create a simple table Excel file."""
    if not OPENPYXL_AVAILABLE:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "SimpleTable"

    # Headers (bold)
    headers = ["Name", "Age", "City", "Salary"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Data
    data = [
        ["Alice Johnson", 28, "New York", 75000],
        ["Bob Smith", 34, "San Francisco", 90000],
        ["Carol Davis", 29, "Chicago", 68000],
        ["David Wilson", 42, "Austin", 85000],
        ["Eva Martinez", 31, "Seattle", 82000],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    print(f"Created: {filepath}")


def create_large_table_xlsx(filepath: Path):
    """Create a large table Excel file to test scaling."""
    if not OPENPYXL_AVAILABLE:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "LargeTable"

    # Headers
    headers = [f"Column_{i}" for i in range(1, 21)]  # 20 columns
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Generate 100 rows of data
    for row in range(2, 102):  # 100 data rows
        for col in range(1, 21):  # 20 columns
            value = f"R{row-1}C{col}"
            ws.cell(row=row, column=col, value=value)

    wb.save(filepath)
    print(f"Created: {filepath}")


def create_complex_table_xlsx(filepath: Path):
    """Create a complex table with multiple regions and formatting."""
    if not OPENPYXL_AVAILABLE:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "ComplexTable"

    # Table 1: Sales Data (A1:D6)
    ws["A1"] = "Sales Report Q1"
    ws["A1"].font = Font(bold=True)

    # Sales headers
    sales_headers = ["Product", "Units Sold", "Price", "Total"]
    for col, header in enumerate(sales_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    # Sales data
    sales_data = [
        ["Widget A", 150, 19.99, "=B3*C3"],
        ["Widget B", 200, 24.99, "=B4*C4"],
        ["Widget C", 100, 15.99, "=B5*C5"],
        ["Widget D", 75, 29.99, "=B6*C6"],
    ]

    for row_idx, row_data in enumerate(sales_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Table 2: Employee Data (F1:I5)
    ws["F1"] = "Employee List"
    ws["F1"].font = Font(bold=True)

    # Employee headers
    emp_headers = ["Name", "Department", "Salary", "Start Date"]
    for col, header in enumerate(emp_headers, 6):  # Starting at column F
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    # Employee data
    emp_data = [
        ["John Doe", "Engineering", 85000, "2023-01-15"],
        ["Jane Smith", "Marketing", 70000, "2023-02-01"],
        ["Mike Johnson", "Sales", 65000, "2023-03-01"],
    ]

    for row_idx, row_data in enumerate(emp_data, 3):
        for col_idx, value in enumerate(row_data, 6):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Table 3: Summary (A8:C10)
    ws["A8"] = "Summary"
    ws["A8"].font = Font(bold=True)

    summary_data = [["Total Products", 4], ["Total Employees", 3]]

    for row_idx, row_data in enumerate(summary_data, 9):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    print(f"Created: {filepath}")


def create_simple_table_csv(filepath: Path):
    """Create a simple CSV file."""
    data = [
        ["Name", "Age", "City", "Salary"],
        ["Alice Johnson", 28, "New York", 75000],
        ["Bob Smith", 34, "San Francisco", 90000],
        ["Carol Davis", 29, "Chicago", 68000],
        ["David Wilson", 42, "Austin", 85000],
        ["Eva Martinez", 31, "Seattle", 82000],
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(data)

    print(f"Created: {filepath}")


def create_large_table_csv(filepath: Path):
    """Create a large CSV file."""
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Headers
        headers = [f"Column_{i}" for i in range(1, 21)]
        writer.writerow(headers)

        # Data rows
        for row in range(1, 101):  # 100 rows
            row_data = [f"R{row}C{col}" for col in range(1, 21)]
            writer.writerow(row_data)

    print(f"Created: {filepath}")


def main():
    """Create all test files."""
    # Ensure the manual test directory exists
    manual_dir = Path(__file__).parent
    manual_dir.mkdir(exist_ok=True)

    print("Creating sample test files for manual vision testing...")

    # Create Excel files (if openpyxl is available)
    if OPENPYXL_AVAILABLE:
        create_simple_table_xlsx(manual_dir / "simple_table.xlsx")
        create_large_table_xlsx(manual_dir / "large_table.xlsx")
        create_complex_table_xlsx(manual_dir / "complex_table.xlsx")
    else:
        print("Skipping Excel files - install openpyxl to create .xlsx files")

    # Create CSV files (always available)
    create_simple_table_csv(manual_dir / "simple_table.csv")
    create_large_table_csv(manual_dir / "large_table.csv")

    print("\nTest files created successfully!")
    print(f"Location: {manual_dir}")
    print("\nYou can now run the tests in WEEK3_TESTING_GUIDE.md")

    if not OPENPYXL_AVAILABLE:
        print("\nTo create Excel test files, install openpyxl:")
        print("pip install openpyxl")


if __name__ == "__main__":
    main()
