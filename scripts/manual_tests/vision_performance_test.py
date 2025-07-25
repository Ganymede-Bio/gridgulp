#!/usr/bin/env python3
"""Performance benchmarks for vision module."""

import asyncio
import sys
import time
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root / "src"))

from gridporter.models import FileInfo, FileType  # noqa: E402
from gridporter.models.sheet_data import CellData, SheetData  # noqa: E402
from gridporter.readers import CalamineReader, ExcelReader  # noqa: E402
from gridporter.vision import BitmapAnalyzer, BitmapGenerator  # noqa: E402
from gridporter.vision.integrated_pipeline import IntegratedVisionPipeline  # noqa: E402
from gridporter.vision.pattern_detector import SparsePatternDetector  # noqa: E402


def create_test_sheet(size: str = "medium") -> SheetData:
    """Create test sheet for benchmarking.

    Args:
        size: "small" (100x50), "medium" (1000x100), or "large" (10000x200)

    Returns:
        SheetData with test content
    """
    sizes = {"small": (100, 50), "medium": (1000, 100), "large": (10000, 200)}

    rows, cols = sizes.get(size, (1000, 100))

    # Create standard SheetData
    sheet = SheetData(name=f"TestSheet_{size}")

    # Create headers
    for col in range(cols):
        sheet[f"{chr(65 + col % 26)}{1}"] = CellData(
            value=f"Column {col + 1}", data_type="text", is_bold=True
        )

    # Add some data (sparse)
    for row in range(1, rows):
        # Add data to ~30% of cells
        for col in range(cols):
            if (row + col) % 3 == 0:
                cell_addr = f"{chr(65 + col % 26)}{row + 1}"
                if row % 10 == 0:
                    # Some formulas
                    sheet[cell_addr] = CellData(
                        value=f"=SUM(A{row}:B{row})", data_type="formula", has_formula=True
                    )
                elif col % 5 == 0:
                    # Some numbers
                    sheet[cell_addr] = CellData(value=row * col, data_type="number")
                else:
                    # Text data
                    sheet[cell_addr] = CellData(value=f"Data_{row}_{col}", data_type="text")

    sheet.max_row = rows - 1
    sheet.max_column = cols - 1

    return sheet


def benchmark_bitmap_generation():
    """Benchmark bitmap generation."""
    print("\n=== Bitmap Generation Benchmarks ===")

    generator = BitmapGenerator()

    for size in ["small", "medium", "large"]:
        print(f"\nTesting {size} sheet:")

        sheet = create_test_sheet(size)

        # Benchmark
        start = time.time()
        img_bytes, metadata = generator.generate(sheet)
        elapsed_time = time.time() - start

        print(f"  Time: {elapsed_time:.3f}s")
        print(f"    Bitmap size: {metadata.width}x{metadata.height}")
        print(f"    File size: {len(img_bytes) / 1024:.1f} KB")


def benchmark_bitmap_analysis():
    """Benchmark bitmap analysis."""
    print("\n=== Bitmap Analysis Benchmarks ===")

    analyzer = BitmapAnalyzer()

    for size in ["small", "medium", "large"]:
        print(f"\nTesting {size} sheet:")

        sheet = create_test_sheet(size)

        # Benchmark
        start = time.time()
        bitmap, metadata = analyzer.generate_binary_bitmap(sheet)
        elapsed_time = time.time() - start

        print(f"  Time: {elapsed_time:.3f}s")
        print(f"    Filled cells: {metadata['filled_cells']}")
        print(f"    Density: {metadata['density']:.1%}")


def benchmark_pattern_detection():
    """Benchmark pattern detection."""
    print("\n=== Pattern Detection Benchmarks ===")

    detector = SparsePatternDetector()

    for size in ["small", "medium"]:  # Skip large for pattern detection
        print(f"\nTesting {size} sheet:")

        sheet = create_test_sheet(size)

        # Benchmark
        start = time.time()
        patterns = detector.detect_patterns(sheet)
        elapsed_time = time.time() - start

        print(f"  Time: {elapsed_time:.3f}s")
        print(f"    Patterns found: {len(patterns)}")


def benchmark_integrated_pipeline():
    """Benchmark the full integrated vision pipeline."""
    print("\n=== Integrated Pipeline Benchmarks ===")

    pipeline = IntegratedVisionPipeline()

    for size in ["small", "medium"]:
        print(f"\nTesting {size} sheet:")

        sheet = create_test_sheet(size)

        # Benchmark
        start = time.time()
        result = pipeline.process_sheet(sheet)
        elapsed_time = time.time() - start

        print(f"  Time: {elapsed_time:.3f}s")
        print(f"    Tables detected: {len(result.detected_tables)}")
        print(f"    Visualization regions: {len(result.visualization_regions)}")


async def benchmark_file_reading():
    """Benchmark reading real Excel files with different readers."""
    print("\n=== Excel File Reading Benchmarks ===")

    # Try to find a test Excel file
    test_files = [
        Path("tests/manual/test_large.xlsx"),
        Path("tests/data/sample.xlsx"),
        Path("sample.xlsx"),
    ]

    test_file = None
    for f in test_files:
        if f.exists():
            test_file = f
            break

    if not test_file:
        print("  No test Excel file found. Creating one...")
        # Create a test file
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # Add headers
        for col in range(1, 51):
            ws.cell(1, col, f"Column {col}").font = openpyxl.styles.Font(bold=True)

        # Add data
        for row in range(2, 1002):
            for col in range(1, 51):
                if (row + col) % 3 == 0:
                    ws.cell(row, col, f"Data_{row}_{col}")

        test_file = Path("test_benchmark.xlsx")
        wb.save(test_file)
        print(f"  Created test file: {test_file}")

    file_info = FileInfo(path=test_file, type=FileType.XLSX, size=test_file.stat().st_size)

    print(f"\nTesting file: {test_file.name} ({file_info.size_mb:.1f} MB)")

    # Benchmark openpyxl reader
    print("\n  Openpyxl Reader:")
    excel_reader = ExcelReader(test_file, file_info)
    start = time.time()
    file_data = await excel_reader.read()
    openpyxl_time = time.time() - start
    print(f"    Time: {openpyxl_time:.3f}s")
    print(f"    Sheets: {len(file_data.sheets)}")
    if file_data.sheets:
        print(f"    First sheet cells: {len(file_data.sheets[0].cells)}")

    # Benchmark Calamine reader
    print("\n  Calamine Reader:")
    try:
        calamine_reader = CalamineReader(test_file, file_info)
        start = time.time()
        file_data_c = await calamine_reader.read()
        calamine_time = time.time() - start
        print(f"    Time: {calamine_time:.3f}s")
        print(f"    Sheets: {len(file_data_c.sheets)}")
        if file_data_c.sheets:
            print(f"    First sheet cells: {len(file_data_c.sheets[0].cells)}")
        print(f"    Speedup: {openpyxl_time / calamine_time:.1f}x")

        # Test direct to Polars
        print("\n  Calamine → Polars DataFrames:")
        start = time.time()
        dfs = calamine_reader.read_to_polars()
        polars_time = time.time() - start
        print(f"    Time: {polars_time:.3f}s")
        print(f"    DataFrames: {len(dfs)}")
        if dfs:
            print(f"    First DataFrame shape: {dfs[0].shape}")
        print(f"    Speedup vs openpyxl: {openpyxl_time / polars_time:.1f}x")

    except ImportError:
        print("    Calamine not available (install python-calamine)")
    except Exception as e:
        print(f"    Error: {e}")

    # Clean up if we created a test file
    if test_file.name == "test_benchmark.xlsx":
        test_file.unlink()


def main():
    """Run all benchmarks."""
    print("GridPorter Vision Module Performance Benchmarks")
    print("=" * 50)
    print("\nPerformance Notes:")
    print("• GridPorter is optimized for table detection, not bulk data processing")
    print("• Vision operations use sparse cell access patterns")
    print("• For data analysis after extraction, convert to Polars/Pandas")
    print("-" * 40)

    # Run synchronous benchmarks
    benchmark_bitmap_generation()
    benchmark_bitmap_analysis()
    benchmark_pattern_detection()
    benchmark_integrated_pipeline()

    # Run async benchmarks
    asyncio.run(benchmark_file_reading())

    print("\n✅ All benchmarks complete!")
    print("\nKey Takeaways:")
    print("• For Excel reading: Use CalamineReader (10-100x faster)")
    print("• For table detection: GridPorter's vision module is efficient")
    print("• For post-processing: Export to Polars/Pandas DataFrames")
    print("• Enable telemetry for production monitoring")


if __name__ == "__main__":
    main()
