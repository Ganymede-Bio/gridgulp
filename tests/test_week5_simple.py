"""Simplified Week 5 Test Suite that works with current implementation.

This test file demonstrates Week 5 functionality:
- Multi-row header detection
- Semantic structure analysis
- Complex table agent
- Feature collection (where supported)
"""

import asyncio
import tempfile
from pathlib import Path

import sys

sys.path.insert(0, "src")

from gridporter import GridPorter
from gridporter.config import Config
from gridporter.models.sheet_data import SheetData, CellData
from gridporter.models.table import TableRange
from gridporter.detectors.multi_header_detector import MultiHeaderDetector
from gridporter.detectors.format_analyzer import SemanticFormatAnalyzer, RowType
from gridporter.agents.complex_table_agent import ComplexTableAgent
from gridporter.telemetry import get_feature_collector
from gridporter.telemetry.feature_store import FeatureStore


def test_multi_row_headers():
    """Test multi-row header detection."""
    print("\n=== Testing Multi-Row Headers ===")

    # Create sheet with multi-row headers
    sheet = SheetData(name="MultiRowHeaders")

    # Row 0: Main categories
    sheet.set_cell(0, 0, CellData(value="Product Info", is_bold=True))
    sheet.set_cell(0, 2, CellData(value="Sales Data", is_bold=True))

    # Row 1: Sub-headers
    sheet.set_cell(1, 0, CellData(value="Name", is_bold=True))
    sheet.set_cell(1, 1, CellData(value="Category", is_bold=True))
    sheet.set_cell(1, 2, CellData(value="Q1", is_bold=True))
    sheet.set_cell(1, 3, CellData(value="Q2", is_bold=True))

    # Data rows
    sheet.set_cell(2, 0, CellData(value="Widget A"))
    sheet.set_cell(2, 1, CellData(value="Hardware"))
    sheet.set_cell(2, 2, CellData(value=100, data_type="number"))
    sheet.set_cell(2, 3, CellData(value=150, data_type="number"))

    # Define table range
    table_range = TableRange(start_row=0, start_col=0, end_row=2, end_col=3)

    # Detect multi-row headers
    detector = MultiHeaderDetector()
    result = detector.detect_multi_row_headers(sheet, table_range)

    if result:
        print(f"✓ Multi-row headers detected: rows {result.start_row} to {result.end_row}")
        print(f"  Confidence: {result.confidence:.2f}")
        print(f"  Column mappings: {len(result.column_mappings)} columns")
        return True
    else:
        print("✗ No multi-row headers detected")
        return False


def test_semantic_structure():
    """Test semantic structure analysis."""
    print("\n=== Testing Semantic Structure Analysis ===")

    # Create financial report
    sheet = SheetData(name="FinancialReport")

    # Headers
    sheet.set_cell(0, 0, CellData(value="Account", is_bold=True))
    sheet.set_cell(0, 1, CellData(value="Q1", is_bold=True))
    sheet.set_cell(0, 2, CellData(value="Q2", is_bold=True))

    # Section: Revenue
    sheet.set_cell(1, 0, CellData(value="Revenue", is_bold=True, background_color="#E0E0E0"))

    # Revenue items
    sheet.set_cell(2, 0, CellData(value="  Product Sales", indentation_level=1))
    sheet.set_cell(2, 1, CellData(value=1000, data_type="number"))
    sheet.set_cell(2, 2, CellData(value=1200, data_type="number"))

    sheet.set_cell(3, 0, CellData(value="  Services", indentation_level=1))
    sheet.set_cell(3, 1, CellData(value=500, data_type="number"))
    sheet.set_cell(3, 2, CellData(value=600, data_type="number"))

    # Subtotal
    sheet.set_cell(4, 0, CellData(value="Total Revenue", is_bold=True))
    sheet.set_cell(4, 1, CellData(value=1500, data_type="number", is_bold=True))
    sheet.set_cell(4, 2, CellData(value=1800, data_type="number", is_bold=True))

    # Analyze structure
    analyzer = SemanticFormatAnalyzer()
    table_range = TableRange(start_row=0, start_col=0, end_row=4, end_col=2)
    structure = analyzer.analyze_table_structure(sheet, table_range, header_rows=1)

    print("✓ Structure analyzed:")
    print(f"  Has subtotals: {structure.has_subtotals}")
    print(f"  Sections found: {len(structure.sections)}")

    # Check row types
    row_types = [row.row_type for row in structure.semantic_rows[:5]]
    print(f"  Row types: {[rt.value for rt in row_types]}")

    # Check if we have the expected structure
    has_total_row = any(row.row_type == RowType.TOTAL for row in structure.semantic_rows)
    has_section = len(structure.sections) > 0

    return has_total_row and has_section


async def test_complex_table_agent():
    """Test complex table detection agent."""
    print("\n=== Testing Complex Table Agent ===")

    config = Config(use_vision=False, suggest_names=False, confidence_threshold=0.5)

    agent = ComplexTableAgent(config)

    # Create test sheet
    sheet = SheetData(name="TestSheet")
    sheet.set_cell(0, 0, CellData(value="Name", is_bold=True))
    sheet.set_cell(0, 1, CellData(value="Value", is_bold=True))
    sheet.set_cell(1, 0, CellData(value="Item1"))
    sheet.set_cell(1, 1, CellData(value=100, data_type="number"))
    sheet.set_cell(2, 0, CellData(value="Item2"))
    sheet.set_cell(2, 1, CellData(value=200, data_type="number"))

    result = await agent.detect_complex_tables(sheet)

    print("✓ Complex detection completed:")
    print(f"  Tables found: {len(result.tables)}")
    print(f"  Confidence: {result.confidence:.2f}")

    if result.tables:
        table = result.tables[0]
        print(f"  First table range: {table.range.excel_range}")
        print(f"  Has headers: {table.has_headers}")
        return True

    return False


async def test_gridporter_integration():
    """Test full GridPorter integration."""
    print("\n=== Testing GridPorter Integration ===")

    import openpyxl

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        test_file = Path(tmp.name)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # Add headers
        ws["A1"] = "Product"
        ws["B1"] = "Sales"

        # Make headers bold
        for cell in ["A1", "B1"]:
            ws[cell].font = openpyxl.styles.Font(bold=True)

        # Add data
        ws["A2"] = "Widget A"
        ws["B2"] = 100
        ws["A3"] = "Widget B"
        ws["B3"] = 200

        wb.save(test_file)
        wb.close()

    try:
        # Initialize GridPorter
        gridporter = GridPorter(use_vision=False, suggest_names=False, confidence_threshold=0.5)

        # Detect tables
        result = await gridporter.detect_tables(test_file)

        print("✓ GridPorter detection completed:")
        print(f"  File processed: {test_file.name}")
        print(f"  Sheets: {len(result.sheets)}")
        print(f"  Total tables: {result.metadata['total_tables']}")

        return result.metadata["total_tables"] > 0

    finally:
        if test_file.exists():
            test_file.unlink()


def test_feature_collection():
    """Test feature collection (if enabled)."""
    print("\n=== Testing Feature Collection ===")

    with tempfile.TemporaryDirectory() as tmpdir:
        feature_db = Path(tmpdir) / "test_features.db"

        # Initialize feature collector
        feature_collector = get_feature_collector()
        feature_collector.initialize(enabled=True, db_path=str(feature_db))

        if feature_collector.enabled:
            print("✓ Feature collection enabled")

            # Check we can get statistics
            stats = feature_collector.get_summary_statistics()
            if stats:
                print(f"  Total records: {stats.get('total_records', 0)}")

            return True
        else:
            print("✗ Feature collection not enabled")
            return False


async def main():
    """Run all tests."""
    print("=" * 60)
    print("WEEK 5 SIMPLIFIED TEST SUITE")
    print("=" * 60)

    results = []

    # Run tests
    results.append(("Multi-Row Headers", test_multi_row_headers()))
    results.append(("Semantic Structure", test_semantic_structure()))
    results.append(("Complex Table Agent", await test_complex_table_agent()))
    results.append(("GridPorter Integration", await test_gridporter_integration()))
    results.append(("Feature Collection", test_feature_collection()))

    # Summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "✓ PASSED" if result else "✗ FAILED"
        print(f"{test_name}: {status}")

    print(f"\nTotal: {passed}/{total} passed ({passed/total*100:.0f}%)")

    if passed == total:
        print("\n✅ ALL TESTS PASSED!")
    else:
        print(f"\n❌ {total - passed} tests failed")

    return passed == total


if __name__ == "__main__":
    success = asyncio.run(main())
    import sys

    sys.exit(0 if success else 1)
