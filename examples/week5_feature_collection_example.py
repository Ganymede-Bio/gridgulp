"""Example of using GridPorter Week 5 features with feature collection enabled.

This example demonstrates:
- Complex table detection with multi-row headers
- Semantic structure analysis (sections, subtotals)
- Feature collection and analysis
- Exporting collected features for analysis
"""

import asyncio
from pathlib import Path

import openpyxl
from gridporter import GridPorter
from gridporter.telemetry import get_feature_collector
from gridporter.telemetry.feature_store import FeatureStore


async def create_complex_spreadsheet(file_path: Path):
    """Create a complex Excel file with multi-row headers and semantic structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Q4 Sales Report"

    # Multi-level headers
    # Level 1: Main categories
    ws["A1"] = "Region"
    ws.merge_cells("A1:A3")
    ws["B1"] = "Product Performance"
    ws.merge_cells("B1:G1")
    ws["H1"] = "Customer Metrics"
    ws.merge_cells("H1:J1")

    # Level 2: Subcategories
    ws["B2"] = "Hardware"
    ws.merge_cells("B2:D2")
    ws["E2"] = "Software"
    ws.merge_cells("E2:G2")
    ws["H2"] = "Satisfaction"
    ws["I2"] = "Retention"
    ws["J2"] = "Growth"

    # Level 3: Metrics
    headers_row3 = [
        "Units",
        "Revenue",
        "Margin",
        "Units",
        "Revenue",
        "Margin",
        "Score",
        "Rate %",
        "New %",
    ]
    for i, header in enumerate(headers_row3):
        ws.cell(row=3, column=i + 2, value=header)

    # Apply bold formatting to headers
    for row in range(1, 4):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = openpyxl.styles.Font(bold=True)

    # Data sections with subtotals
    regions = {
        "North America": [
            ["USA", 500, 50000, 0.25, 200, 40000, 0.45, 4.5, 92, 15],
            ["Canada", 150, 15000, 0.22, 80, 16000, 0.42, 4.3, 89, 12],
            ["Mexico", 100, 8000, 0.20, 50, 8000, 0.40, 4.1, 85, 18],
        ],
        "Europe": [
            ["UK", 300, 35000, 0.28, 150, 35000, 0.48, 4.6, 93, 14],
            ["Germany", 400, 45000, 0.26, 180, 40000, 0.46, 4.7, 94, 11],
            ["France", 250, 25000, 0.24, 120, 25000, 0.44, 4.4, 91, 13],
        ],
        "Asia Pacific": [
            ["Japan", 600, 70000, 0.30, 300, 65000, 0.50, 4.8, 95, 20],
            ["China", 800, 75000, 0.22, 400, 70000, 0.42, 4.2, 88, 25],
            ["Australia", 200, 22000, 0.27, 100, 20000, 0.47, 4.5, 92, 16],
        ],
    }

    current_row = 4
    for region, countries in regions.items():
        # Section header
        ws.cell(row=current_row, column=1, value=region)
        cell = ws.cell(row=current_row, column=1)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )
        current_row += 1

        # Country data
        for country_data in countries:
            country_name = country_data[0]
            ws.cell(row=current_row, column=1, value=f"  {country_name}")  # Indented
            for col, value in enumerate(country_data[1:], 2):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

        # Subtotal row
        ws.cell(row=current_row, column=1, value=f"Total {region}")
        cell = ws.cell(row=current_row, column=1)
        cell.font = openpyxl.styles.Font(bold=True)

        # Calculate subtotals (simplified - just sum numeric columns)
        for col in range(2, 11):
            if col in [2, 3, 5, 6]:  # Sum columns
                total = sum(
                    row[col - 1] for row in countries if isinstance(row[col - 1], int | float)
                )
                ws.cell(row=current_row, column=col, value=total)
                ws.cell(row=current_row, column=col).font = openpyxl.styles.Font(bold=True)
            elif col in [4, 7]:  # Average columns
                avg = sum(row[col - 1] for row in countries) / len(countries)
                ws.cell(row=current_row, column=col, value=round(avg, 2))
                ws.cell(row=current_row, column=col).font = openpyxl.styles.Font(bold=True)

        current_row += 1
        # Blank row between sections
        current_row += 1

    # Grand total row
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    grand_total_cell = ws.cell(row=current_row, column=1)
    grand_total_cell.font = openpyxl.styles.Font(bold=True, color="000080")
    grand_total_cell.fill = openpyxl.styles.PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )

    # Apply alternating row colors to data rows
    for row in range(5, current_row):
        if ws.cell(row=row, column=1).value and not ws.cell(row=row, column=1).font.bold:
            fill_color = "F5F5F5" if row % 2 == 0 else "FFFFFF"
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    ws.column_dimensions["A"].width = 15

    wb.save(file_path)
    print(f"Created complex spreadsheet: {file_path}")


async def analyze_with_features():
    """Demonstrate Week 5 complex table detection with feature collection."""
    # Create test file
    test_file = Path("complex_sales_report.xlsx")
    await create_complex_spreadsheet(test_file)

    try:
        # Initialize GridPorter with feature collection enabled
        gridporter = GridPorter(
            use_vision=False,
            suggest_names=True,
            enable_feature_collection=True,
            feature_db_path="week5_features.db",
        )

        print("\nAnalyzing complex spreadsheet with feature collection enabled...")

        # Detect tables
        result = await gridporter.detect_tables(test_file)

        # Display results
        print("\nDetection Results:")
        print(f"  File: {result.file_info.path}")
        print(f"  Detection time: {result.detection_time:.2f}s")
        print(f"  Total tables found: {result.metadata.get('total_tables', 0)}")

        for sheet in result.sheets:
            print(f"\n  Sheet: {sheet.name}")
            for i, table in enumerate(sheet.tables):
                print(f"\n    Table {i+1}:")
                print(f"      Range: {table.range}")
                print(f"      Confidence: {table.confidence:.2f}")
                print(f"      Detection method: {table.detection_method}")

                if table.header_info:
                    print(
                        f"      Headers: {'Multi-row' if table.header_info.is_multi_row else 'Single-row'}"
                    )
                    print(f"      Header rows: {table.header_info.row_count}")

                if table.semantic_structure:
                    print("      Semantic features:")
                    print(
                        f"        Has sections: {table.semantic_structure.get('has_sections', False)}"
                    )
                    print(
                        f"        Section count: {table.semantic_structure.get('section_count', 0)}"
                    )
                    print(
                        f"        Has subtotals: {table.semantic_structure.get('has_subtotals', False)}"
                    )
                    print(
                        f"        Has grand total: {table.semantic_structure.get('has_grand_total', False)}"
                    )

                if table.format_preservation:
                    print("      Format features:")
                    print(
                        f"        Alternating rows: {table.format_preservation.get('alternating_rows', False)}"
                    )
                    print(
                        f"        Preserved blank rows: {table.format_preservation.get('blank_rows', [])}"
                    )

        # Analyze collected features
        print("\n" + "=" * 60)
        print("FEATURE COLLECTION ANALYSIS")
        print("=" * 60)

        feature_collector = get_feature_collector()
        if feature_collector.enabled:
            # Get statistics
            stats = feature_collector.get_summary_statistics()
            if stats:
                print("\nOverall Statistics:")
                print(f"  Total detections: {stats['total_count']}")
                print(f"  Success rate: {stats['success_rate']:.1%}")
                print(f"  Average confidence: {stats.get('avg_confidence', 0):.2f}")
                print(f"  Detection methods used: {', '.join(stats.get('detection_methods', []))}")

                # Method-specific stats
                if "method_stats" in stats:
                    complex_stats = stats["method_stats"].get("complex_detection", {})
                    if complex_stats:
                        print("\nComplex Detection Statistics:")
                        print(f"  Count: {complex_stats.get('count', 0)}")
                        print(
                            f"  Average processing time: {complex_stats.get('avg_processing_time', 0):.1f}ms"
                        )
                        print(
                            f"  Multi-row header detections: {complex_stats.get('multi_row_header_count', 0)}"
                        )

            # Export features for further analysis
            export_path = "week5_features_export.csv"
            feature_collector.export_features(export_path, format="csv")
            print(f"\nFeatures exported to: {export_path}")

            # Query specific features
            store = FeatureStore("week5_features.db")
            recent_features = store.query_features(limit=5)

            if recent_features:
                print("\nRecent Detection Features:")
                for feat in recent_features:
                    print(f"\n  File: {Path(feat.file_path).name}")
                    print(f"    Method: {feat.detection_method}")
                    print(f"    Confidence: {feat.confidence:.2f}")
                    print(f"    Header rows: {feat.header_row_count}")
                    print(f"    Has multi-row headers: {feat.has_multi_row_headers}")
                    print(f"    Has subtotals: {feat.has_subtotals}")
                    print(f"    Section count: {feat.section_count}")
                    print(f"    Processing time: {feat.processing_time_ms}ms")

            store.close()

        await gridporter.close()

    finally:
        # Cleanup
        if test_file.exists():
            test_file.unlink()
            print(f"\nCleaned up test file: {test_file}")


if __name__ == "__main__":
    print("GridPorter Week 5 Feature Collection Example")
    print("=" * 50)
    print("This example demonstrates:")
    print("- Complex table detection with multi-row headers")
    print("- Semantic structure analysis")
    print("- Feature collection and telemetry")
    print("- Exporting features for analysis")
    print("=" * 50)

    asyncio.run(analyze_with_features())
